#! /usr/bin/python3

import os
import re
import sys
import warnings
import shutil
import json
import argparse

warnings.filterwarnings("ignore")
#may not need this
#sys.path.append('/Users/ronnieperez/Desktop/Stuff/Programming/python/modules')

import openpyxl

warnings.simplefilter(action='ignore', category=UserWarning)

parser = argparse.ArgumentParser()

parser.add_argument('--xlf',  type=str, required=True)
parser.add_argument('--lcol', type=str, required=True)

args = parser.parse_args()

xlinf   = args.xlf
lastCol = int(args.lcol)

firstCol = 2       #first column of worksheet is 2 (that data is written to)
firstRow = 5       #first row    of worksheet is 5 (tat data we need is in)

jsonf = "fastchannel.json"

section_list = ["filename", "format", "subtitlelanguage", "subgenre", "programversion", "sccfilename", "housenumber"]

#define alphabet dictionary
letterDictionary = {
    'A' : 1,
    'B' : 2,
    'C' : 3,
    'D' : 4,
    'E' : 5,
    'F' : 6,
    'G' : 7,
    'H' : 8,
    'I' : 9,
    'J' : 10,
    'K' : 11,
    'L' : 12,
    'M' : 13,
    'N' : 14,
    'O' : 15,
    'P' : 16,
    'Q' : 17,
    'R' : 18,
    'S' : 19,
    'T' : 20,
    'U' : 21,
    'V' : 22,
    'W' : 23,
    'X' : 24,
    'Y' : 25,
    'Z' : 26
}

#converts the xl cells to their corresponding number
def getColumnNumber(col):

    sum = 0
    power = len(col) - 1

    for l in col:
        l = l.upper()
        n = letterDictionary[l]
        sum += n * (26 ** power)
        power -= 1

    return sum


"""
#get input file from user
def getInputFile():

    print("")
    while True:
        inf = input("Give me your input file: ")
        if os.path.exists(inf):
            break
        else:
            print("   Does not exist")

    return inf
"""

def getXLModifyCols(ws,endRow,endCol):

    d = {}    #define an empty dictionary

    rowCount = firstRow  # a counter to keep track of the row I am on
    colCount = firstCol  #  counter to keep track of the col I am on

    #read the contents
    for category in ws.iter_rows(min_row=firstRow,min_col=firstCol,max_row=endRow,max_col=endCol,values_only=True):

        """
        """
        print("\nAnalyzing xl sheet.....")


    for i in range (0,len(category)):
        #print(i,category[i])
        """
        """
        word = category[i]
        word = word.lower()
        word = word.replace("\n","")
        word = word.replace(" ","")
        word = word.replace("#","num")
        word = word.replace("-","")
        word = word.replace(".","")
        word = word.replace("(noextension)","")
        #print(i,word)
        n = firstCol + i  #the column number is offset by 2, becuase first data is in column B which is 2
        d[word] = n       #add the word and column to the hash
        """
        """

    #lets do a check to see if all of these were found
    rat = 0
    for section in section_list:

        if not (section in d):
            print("\n  ",section,"could not be found in xl file:")
            rat += 1

    if rat > 0:
        print("")
        sys.exit(1)

    return d

"""
#get the last column from the user, as a integer
def getLastCol():

    print("")
    while True:
        col = input("Give me the last column in your sheet ie [AA]: ")
        m = re.match('^[a-zA-Z]+$',col)
        if m:
            col = col.upper()
            #print(col)
            break
        else:
            print("  Invlaid Format")

    return getColumnNumber(col)
"""

#make a backup of the xl file becuase we will modify the original
def backupWB(xlinf):

    newf = xlinf + ".backup"
    print("\nMaking backup file")
    try:
        shutil.copyfile(xlinf,newf)
    except:
        print("Error copying backup file")

def countNumEpisodes():

    n = 0;
    f  = open(jsonf,"r")
    for l in f.readlines():
        l = l.rstrip()

        pattern = '"housenumber"'

        m = re.search(pattern,l)
        if m:
            n += 1;
    return n

#read the json file
def readJSON():

    jsonf = "fastchannel.json"

    # Opening JSON file
    f = open(jsonf)
    data = json.load(f)
    f.close()

    return data

#xlinf   = getInputFile()
#lastCol = getLastCol(lcnum)   #lastCol returned as integer

workbook  = openpyxl.load_workbook(filename=xlinf)  #set the load_workbook
worksheet = workbook["1. Master Metadata"]          #set the name of the worksheet to read (later have user enter)
categorydict = getXLModifyCols(worksheet,5,lastCol) #got latCol from user, hardcoding 5 right now
backupWB(xlinf) #copy the workbook

numEpisodes = countNumEpisodes()
#print("Number of episodes :",numEpisodes)


showdata = readJSON()

startcol = firstCol
startrow = firstRow + 1

print("")
n = 0
for rn in range(startrow,(startrow + numEpisodes)):

    for sec in section_list:
        cn  = categorydict[sec]          #column number
        val = showdata['season'][n][sec] #

        print("row",rn,":",sec,":",val)
        worksheet.cell(row=rn,column=cn).value = val
        print("===========")
    n += 1

workbook.save(xlinf)
workbook.close()
